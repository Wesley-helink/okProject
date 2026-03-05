#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
HMA Medical Management System - Man-day Estimation Tool
Analyzes features from Excel file and generates effort estimates
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# Configuration
EXCEL_INPUT = 'HMA_醫務管理/HMA_醫務管理/第二階段_功能與程式清單.xlsx'
EXCEL_OUTPUT = 'HMA_開發人天評估表_含維護.xlsx'

# Estimation parameters (man-days per complexity)
COMPLEXITY_ESTIMATES = {
    '簡單': 2,      # Simple: 2 days
    '中等': 5,      # Medium: 5 days
    '複雜': 10,     # Complex: 10 days
    '非常複雜': 15  # Very Complex: 15 days
}

# Hourly rate (NT$/hour) - adjust as needed
HOURLY_RATE = 1200
HOURS_PER_DAY = 8

# Additional work items
ADDITIONAL_ITEMS = [
    {'name': '排除弱點掃描中及高風險的修正', 'man_days': 10, 'category': '安全性'},
    {'name': 'SA 文件 (System Analysis)', 'man_days': 8, 'category': '文件'},
    {'name': 'SD 文件 (System Design)', 'man_days': 10, 'category': '文件'},
    {'name': '測試報告文件', 'man_days': 6, 'category': '文件'},
    {'name': '操作手冊文件', 'man_days': 8, 'category': '文件'},
    {'name': '爬網(程式)優化', 'man_days': 12, 'category': '功能開發'},
    {'name': '保固1年', 'man_days': 30, 'category': '維護'},
    {'name': '小修改維護1年', 'man_days': 40, 'category': '維護'}
]

def estimate_complexity(feature_name, feature_type='畫面作業'):
    """
    Estimate complexity based on feature name and type
    Returns: (complexity_level, man_days)
    """
    name_lower = feature_name.lower() if feature_name else ''
    
    # Complex features (keywords)
    if any(keyword in feature_name for keyword in ['批次', '整批', '匯入', '轉入', '上傳', '整合型']):
        return '複雜', 10
    
    # Medium-complex features
    if any(keyword in feature_name for keyword in ['申請', '作業', '照護計畫', '試辦']):
        return '中等', 5
    
    # Report operations are generally simpler
    if feature_type == '報表作業' or '列印' in feature_name or '查詢' in feature_name:
        return '簡單', 3
    
    # Default to medium
    return '中等', 5

def read_excel_data():
    """Read and parse the Excel file"""
    print(f"Reading Excel file: {EXCEL_INPUT}")

    wb = openpyxl.load_workbook(EXCEL_INPUT, data_only=True)

    # Try to find the correct sheet
    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")

    # Look for sheet with features
    target_sheet = None
    for name in ['功能程式清單', 'Sheet1', sheet_names[0]]:
        if name in sheet_names:
            target_sheet = wb[name]
            print(f"Using sheet: {name}")
            break

    if not target_sheet:
        target_sheet = wb.active
        print(f"Using active sheet: {target_sheet.title}")

    # Read data
    features = []

    # Header is at row 2: ['模組', '作業(選單)名稱', '功能名稱', '按鍵']
    # Data starts from row 3
    # Columns: B=模組, C=作業(選單)名稱, D=功能名稱, E=按鍵

    print("Parsing feature data...")
    data_start_row = 3

    # Read feature data
    for row in target_sheet.iter_rows(min_row=data_start_row, values_only=True):
        # Column D (index 3) contains feature name
        if row and len(row) > 3 and row[3]:  # Has feature name
            feature_name = str(row[3]).strip()
            if feature_name and feature_name != '功能名稱':
                features.append({
                    'module': str(row[1]) if row[1] else '',
                    'menu': str(row[2]) if row[2] else '',
                    'feature_name': feature_name,
                    'buttons': str(row[4]) if len(row) > 4 and row[4] else ''
                })

    print(f"Found {len(features)} features")
    return features

def categorize_feature(feature_name):
    """Categorize feature based on name - more specific matching first"""
    if not feature_name:
        return '其他'

    # More specific categories first
    if any(kw in feature_name for kw in ['印鑑', '合約', '文件', '拍照']):
        return '文件管理'
    elif any(kw in feature_name for kw in ['試辦', '計畫申請', '照護計畫']):
        return '申請作業'
    elif any(kw in feature_name for kw in ['醫事人員', '人員', '專科證書', '資格', '支援']):
        return '人員管理'
    elif any(kw in feature_name for kw in ['科別', '服務項目', '特殊設備', '病床', '窗口', '聯絡人']):
        return '基本資料維護'
    elif any(kw in feature_name for kw in ['審核', '審查']):
        return '審核作業'
    elif any(kw in feature_name for kw in ['列印', '報表']) and 'R' in feature_name:
        return '報表作業'
    elif any(kw in feature_name for kw in ['查詢', '統計']) and '審核' not in feature_name:
        return '查詢作業'
    elif any(kw in feature_name for kw in ['建檔', '作業', '機構']) and '審核' not in feature_name:
        return '機構管理'
    elif any(kw in feature_name for kw in ['批次', '轉入', '匯入', '上傳', '下載']):
        return '批次處理'
    else:
        return '其他'

def create_estimation_report(features):
    """Create the estimation report Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "開發人天評估"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['序號', '功能名稱', '程式代號', '分類', '複雜度', '預估人天', '開發費用(NT$)', '備註']
    ws.append(headers)
    
    # Style header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Process features
    row_num = 2
    category_totals = {}
    total_days = 0
    total_cost = 0
    
    # Group features by category
    categorized_features = {}
    for idx, feature in enumerate(features, 1):
        # Extract feature info from our parsed structure
        feature_name = feature.get('feature_name', '')

        if not feature_name:
            continue

        feature_name = str(feature_name).strip()

        # Extract program code from feature name (usually in parentheses)
        program_code = ''
        if '(' in feature_name and ')' in feature_name:
            start = feature_name.rfind('(')
            end = feature_name.rfind(')')
            program_code = feature_name[start+1:end]

        if not program_code:
            program_code = f"FUNC{idx:03d}"

        # Determine feature type
        feature_type = '報表作業' if 'R' in program_code else '畫面作業'

        # Estimate complexity and man-days
        complexity, man_days = estimate_complexity(feature_name, feature_type)

        # Categorize
        category = categorize_feature(feature_name)

        # Calculate cost
        cost = man_days * HOURS_PER_DAY * HOURLY_RATE

        # Group by category
        if category not in categorized_features:
            categorized_features[category] = []

        categorized_features[category].append({
            'index': idx,
            'name': feature_name,
            'code': program_code,
            'complexity': complexity,
            'man_days': man_days,
            'cost': cost
        })
    
    # Write categorized features
    for category in sorted(categorized_features.keys()):
        # Category header
        ws.cell(row=row_num, column=1, value=category)
        ws.cell(row=row_num, column=1).font = category_font
        ws.cell(row=row_num, column=1).fill = category_fill
        ws.merge_cells(f'A{row_num}:H{row_num}')
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = border
        
        row_num += 1
        
        category_days = 0
        category_cost = 0
        
        # Features in this category
        for feat in categorized_features[category]:
            ws.append([
                feat['index'],
                feat['name'],
                feat['code'],
                category,
                feat['complexity'],
                feat['man_days'],
                feat['cost'],
                ''
            ])
            
            # Style data row
            for col in range(1, 9):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                if col == 1:  # Index
                    cell.alignment = Alignment(horizontal='center')
                elif col in [6, 7]:  # Numbers
                    cell.alignment = Alignment(horizontal='right')
                    if col == 7:  # Cost
                        cell.number_format = '#,##0'
            
            category_days += feat['man_days']
            category_cost += feat['cost']
            row_num += 1
        
        # Category subtotal
        ws.append(['', f'{category} 小計', '', '', '', category_days, category_cost, ''])
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.border = border
            if col in [6, 7]:
                cell.alignment = Alignment(horizontal='right')
                if col == 7:
                    cell.number_format = '#,##0'
        
        total_days += category_days
        total_cost += category_cost
        row_num += 1
    
    # Add empty row
    row_num += 1
    
    # Add additional work items
    ws.cell(row=row_num, column=1, value='其他工作項目')
    ws.cell(row=row_num, column=1).font = category_font
    ws.cell(row=row_num, column=1).fill = category_fill
    ws.merge_cells(f'A{row_num}:H{row_num}')
    for col in range(1, 9):
        ws.cell(row=row_num, column=col).border = border
    row_num += 1
    
    additional_days = 0
    additional_cost = 0
    
    for idx, item in enumerate(ADDITIONAL_ITEMS, 1):
        man_days = item['man_days']
        cost = man_days * HOURS_PER_DAY * HOURLY_RATE
        
        ws.append([
            f'A{idx}',
            item['name'],
            '-',
            item['category'],
            '-',
            man_days,
            cost,
            item.get('note', '')
        ])
        
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
            elif col in [6, 7]:
                cell.alignment = Alignment(horizontal='right')
                if col == 7:
                    cell.number_format = '#,##0'
        
        additional_days += man_days
        additional_cost += cost
        row_num += 1
    
    # Additional items subtotal
    ws.append(['', '其他工作項目 小計', '', '', '', additional_days, additional_cost, ''])
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = border
        if col in [6, 7]:
            cell.alignment = Alignment(horizontal='right')
            if col == 7:
                cell.number_format = '#,##0'
    row_num += 1
    
    # Grand total
    row_num += 1
    grand_total_days = total_days + additional_days
    grand_total_cost = total_cost + additional_cost
    
    ws.append(['', '總計 (GRAND TOTAL)', '', '', '', grand_total_days, grand_total_cost, ''])
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        if col in [6, 7]:
            cell.alignment = Alignment(horizontal='right')
            if col == 7:
                cell.number_format = '#,##0'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30
    
    # Add summary sheet
    ws_summary = wb.create_sheet("專案摘要")
    
    summary_data = [
        ['HMA 醫務管理系統 - 第二階段開發評估', ''],
        ['', ''],
        ['評估日期', datetime.now().strftime('%Y-%m-%d')],
        ['', ''],
        ['項目', '數值'],
        ['功能開發人天', total_days],
        ['其他工作項目人天', additional_days],
        ['總人天', grand_total_days],
        ['', ''],
        ['每日工時', HOURS_PER_DAY],
        ['時薪 (NT$)', HOURLY_RATE],
        ['', ''],
        ['功能開發費用', total_cost],
        ['其他工作項目費用', additional_cost],
        ['總開發費用 (NT$)', grand_total_cost],
        ['', ''],
        ['預估工期 (以1人計算)', f'{grand_total_days} 天'],
        ['預估工期 (以2人計算)', f'{grand_total_days/2:.1f} 天'],
        ['預估工期 (以3人計算)', f'{grand_total_days/3:.1f} 天'],
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Style summary sheet
    ws_summary['A1'].font = Font(bold=True, size=14, color="366092")
    ws_summary.merge_cells('A1:B1')
    
    for row in range(5, 20):
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        if row in [5, 8, 15]:
            ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        if row in [8, 15]:
            ws_summary.cell(row=row, column=2).font = Font(bold=True, size=12)
            ws_summary.cell(row=row, column=2).number_format = '#,##0'
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Save
    wb.save(EXCEL_OUTPUT)
    print(f"\nEstimation report created: {EXCEL_OUTPUT}")
    print(f"Total man-days: {grand_total_days}")
    print(f"Total cost: NT$ {grand_total_cost:,}")
    
    return grand_total_days, grand_total_cost

def main():
    """Main execution"""
    print("=" * 60)
    print("HMA Medical Management System - Estimation Tool")
    print("=" * 60)
    print()

    try:
        # Read Excel data
        features = read_excel_data()

        if not features:
            print("Warning: No features found in Excel file")
            print("Creating estimation with additional work items only...")

        # Create estimation report
        total_days, total_cost = create_estimation_report(features)

        print()
        print("=" * 60)
        print("Report generation completed!")
        print(f"Output file: {EXCEL_OUTPUT}")
        print("=" * 60)

    except FileNotFoundError:
        print(f"Error: Could not find input file: {EXCEL_INPUT}")
        print("Please ensure the file exists in the correct location")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
