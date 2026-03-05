#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""顯示更新後的評估摘要 (含保固維護)"""

import openpyxl

wb = openpyxl.load_workbook('HMA_開發人天評估表_含維護.xlsx')
ws = wb['開發人天評估']

print('\n' + '='*90)
print('📊 HMA 醫務管理系統 - 更新後評估報告 (含保固與維護)')
print('='*90)
print('\n🛠️  其他工作項目明細 (8項):\n')
print(f"{'項次':<6} {'工作項目':<40} {'人天':<8} {'費用(NT$)':<15} {'分類':<10}")
print('-'*90)

found_other = False
count = 0
total_days = 0
total_cost = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] == '其他工作項目':
        found_other = True
        continue
    
    if found_other and row[0] and row[0] != '序號':
        if '小計' in str(row[1]):
            print('-'*90)
            print(f"{'小計':<6} {'':<40} {row[5]:<8} {row[6]:>13,}")
            print('='*90)
            break
        elif '總計' not in str(row[1]):
            count += 1
            print(f"{str(row[0]):<6} {row[1]:<40} {row[5]:<8} {row[6]:>13,}  {row[3]:<10}")
            total_days += row[5]
            total_cost += row[6]

# 讀取專案摘要
ws_summary = wb['專案摘要']
print('\n' + '='*90)
print('💰 總體評估:')
print('='*90 + '\n')

for row in ws_summary.iter_rows(min_row=3, max_row=19, values_only=True):
    if row[0] and row[0] != '' and row[0] != '項目':
        if isinstance(row[1], (int, float)):
            if row[1] > 1000:
                print(f"  {row[0]:<30} {row[1]:>15,}")
            else:
                print(f"  {row[0]:<30} {row[1]:>15}")
        else:
            print(f"  {row[0]:<30} {row[1]:>15}")

print('\n' + '='*90)
print('⏱️  建議工期:')
print('='*90 + '\n')
print('  • 功能開發: 1071 人天 (不含保固維護)')
print('  • 保固維護: 124 人天 (保固30天 + 小修改維護40天 + 其他54天)')
print('  • 總計: 1195 人天')
print()
print('  建議團隊規模:')
print('  • 5人團隊開發: 約 214 天 (10.7個月) + 保固維護1年')
print('  • 8人團隊開發: 約 134 天 (6.7個月) + 保固維護1年')
print()
print('='*90)
print('✅ 新檔案已產生: HMA_開發人天評估表_含維護.xlsx')
print('='*90 + '\n')
